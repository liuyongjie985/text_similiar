import openpyxl
import sys
import re
import random

input_path = sys.argv[1]
sheetname = sys.argv[2]
sentence_index = int(sys.argv[3])
classify_index = int(sys.argv[4])
output_path = sys.argv[5]

print("reading input")
wb = openpyxl.load_workbook(input_path)
print("done")

sheet = wb[sheetname]
# 按照规则解析表格
max_row = sheet.max_row
max_column = sheet.max_column + 1

o = open(output_path, "w")

i = 1
while i <= max_row:

    def dealClass(i, target):
        temp_list = []
        idx = i
        while sheet.cell(row=idx, column=classify_index).value == target:
            temp_list.append(sheet.cell(row=idx, column=sentence_index).value)
            idx += 1
        return idx, temp_list


    if sheet.cell(row=i, column=classify_index).value == -1:
        i, result = dealClass(i, -1)

        for x, s1 in enumerate(result):
            rd_idx = random.randint(i, max_row)
            o.write(sheet.cell(row=rd_idx, column=sentence_index).value)
            o.write("\t")
            o.write(s1)
            o.write("\t")
            o.write("0")
            o.write("\n")

        continue

    target = sheet.cell(row=i, column=classify_index).value
    i, result = dealClass(i, target)

    for x, s1 in enumerate(result):
        s1 = re.sub("( |\t|\d|(a-z))+", "", s1)
        if len(s1) <= 8:
            continue
        y = x + 1
        count = 0
        while y < len(result):
            s2 = re.sub("( |\t|\d|(a-z))+", "", result[y])
            if len(s2) <= 8:
                y += 1
                continue
            o.write(s1)
            o.write("\t")
            o.write(s2)
            o.write("\t")
            o.write("1")
            o.write("\n")
            count += 1
            if count >= 1:
                count = 0
                break
            y += 1

print(i)
o.close()
wb.close()
